from selenium.webdriver.support.ui import Select
import pandas
import openpyxl
from time import sleep
from os import path
import funcoes
import xlrd

# Planilha Com a base de dados
wbbase = openpyxl.load_workbook('Alunos PMP.xlsx', read_only=True)
wsbase = wbbase.active
ultimalinha = wsbase.max_row + 1


#Limpar Planilha de saida
arquivo = 'Template_arquivo_saida.xlsx'
funcoes.limpar_Template_Saida(arquivo)
sleep(5)

# Planilha de saida com o resultado
wbsaida = openpyxl.load_workbook('Template_arquivo_saida.xlsx')
wsSaida = wbsaida.active
iniPla_Saida = 3
linsaida = wsSaida.max_row + 1
dir = path.dirname(__file__)
wsSaida = wbsaida.sheetnames

#variaveis utilizadas para armazenar o nome e sobre para a pesquisa no site do pmi
lastname = []
firstname = []
colNome = 1
colEmail = 2
colData = 3
colTurma = 4
colNomePMI = 5
colDataPMI = 6
contnome = 0

print('iniciando o Robô...')
driver = funcoes.abrir_site()
resultado = ''

count = 1
for lin in range(2, ultimalinha):
    nome = str(wsbase.cell(row=lin, column=1).value).split()
    print(f'Consultando o aluno: {wsbase.cell(row=lin, column=1).value}')
    firstname = nome[0]
    c = 1
    while c < len(nome):
        lastname = nome[c]
        if len(nome[c]) > 1:
            if len(nome[c]) == 2 or len(nome[c]) == 3:#Verifica se a quantidade de letras do nome tem 2 ou 3 caracteres
                if c == len(nome)-1: #se tiver 2 caracetres ou 3 e for a ultima posição da lista utiliza como ultimo nome
                    lastname = nome[c]
                elif c+1 < len(nome)-1:
                    lastname = nome[c] + ' ' + nome[c + 1]
                    c = c+1
                else: # se tiver dois ou tres letras e não for o ultimo indice da lista de nome concateno com o proximo indice.
                    lastname = nome[c] + ' ' + nome[c + 1]
                    #prenechendo os campos do site
                    clastname = driver.find_element_by_id('dph_RegistryContent_tbSearch')
                    clastname.clear()
                    clastname.send_keys(lastname)
                    cfirstname = driver.find_element_by_id('dph_RegistryContent_firstNameTextBox')
                    cfirstname.clear()
                    cfirstname.send_keys(firstname)
                    Scredential = Select(driver.find_element_by_id('dph_RegistryContent_credentialDDL'))
                    Scredential.select_by_visible_text('PMP')
                    cbutonok = driver.find_element_by_id('dph_RegistryContent_Button1')
                    cbutonok.click()
                    resultado = driver.find_elements_by_xpath('//table')
                    if resultado:
                        cont = 1
                        datapmi = []
                        #Loop para coletar todos os registros do resultado da pesquisa.
                        while True:
                            strcont = str(cont)
                            if driver.find_elements_by_id(
                                    'dph_RegistryContent_SearchResults_ctl0' + strcont + '_certDateLabel1'):
                                datapmi.append(driver.find_element_by_id(
                                    'dph_RegistryContent_SearchResults_ctl0' + strcont + '_certDateLabel1').text)
                                cont += 1
                            else:
                                break
                        for ano in datapmi:
                            anopmi = ano
                            anobase = str(wsbase.cell(row=lin, column=2).value)
                            anobase = anobase[:4]
                            anobase = int(anobase)
                            cont = datapmi.index(ano) + 1
                            strcont = str(cont)
                            if anobase <= int(anopmi[len(anopmi) - 4:]):
                                nomepmi = str(driver.find_element_by_id('dph_RegistryContent_SearchResults_ctl0' + strcont + '_nameLabel1').text).split()
                                if len(nome) >= 3 and len(nomepmi) >= 3:
                                    contnome = 0
                                    for nb in nome:
                                        if len(nb) == 2 or len(nb) == 3:  # Verifica se a quantidade de letras do nome tem 2 ou 3 caracteres
                                            nb = funcoes.verificar_tam_nome(nb, nome)
                                        for n in nomepmi:
                                            n = funcoes.verificar_tam_nome(n, nomepmi)
                                            if nb == n:
                                                contnome += 1
                                    if contnome >= 3:
                                        wsSaida.cell(row=linsaida, column=colNome).value = wsbase.cell(row=lin,
                                                                                                       column=1).value
                                        wsSaida.cell(row=linsaida, column=colData).value = wsbase.cell(row=lin,
                                                                                                       column=2).value
                                        wsSaida.cell(row=linsaida, column=colTurma).value = wsbase.cell(row=lin, column=4).value
                                        wsSaida.cell(row=linsaida, column=colNomePMI).value = driver.find_element_by_id(
                                            'dph_RegistryContent_SearchResults_ctl0' + strcont + '_nameLabel1').text
                                        wsSaida.cell(row=linsaida, column=colDataPMI).value = (datapmi[cont - 1])

                                        wbsaida.save('Template_arquivo_saida.xlsx')
                                        linsaida += 1
                                        break
                                else:
                                    if nome[0] == nomepmi[0]:
                                        contnome = 0
                                        for nb in nome:
                                            if len(nb) == 2 or len(
                                                    nb) == 3:  # Verifica se a quantidade de letras do nome tem 2 ou 3 caracteres
                                                nb = funcoes.verificar_tam_nome(nb, nome)
                                            for n in nomepmi:
                                                n = funcoes.verificar_tam_nome(n, nomepmi)
                                                if nb == n:
                                                    contnome += 1
                                        if contnome >= 2:
                                            wsSaida.cell(row=linsaida, column=colNome).value = wsbase.cell(row=lin, column=1).value
                                            wsSaida.cell(row=linsaida, column=colData).value = wsbase.cell(row=lin, column=2).value
                                            wsSaida.cell(row=linsaida, column=colTurma).value = wsbase.cell(row=lin, column=4).value
                                            wsSaida.cell(row=linsaida, column=colNomePMI).value = driver.find_element_by_id(
                                                'dph_RegistryContent_SearchResults_ctl0' + strcont + '_nameLabel1').text
                                            wsSaida.cell(row=linsaida, column=colDataPMI).value = (datapmi[cont - 1])
                                            wbsaida.save('Template_arquivo_saida.xlsx')
                                            linsaida += 1
                                            break

                    c+= 1
                    sleep(1)
                    break

            clastname = driver.find_element_by_id('dph_RegistryContent_tbSearch')
            clastname.clear()
            clastname.send_keys(lastname)
            cfirstname = driver.find_element_by_id('dph_RegistryContent_firstNameTextBox')
            cfirstname.clear()
            cfirstname.send_keys(firstname)
            Scredential = Select(driver.find_element_by_id('dph_RegistryContent_credentialDDL'))
            Scredential.select_by_visible_text('PMP')
            cbutonok = driver.find_element_by_id('dph_RegistryContent_Button1')
            cbutonok.click()
            resultado = driver.find_elements_by_xpath('//table')
            if resultado:
                cont = 1
                datapmi = []
                while True:
                    strcont = str(cont)
                    if driver.find_elements_by_id(
                            'dph_RegistryContent_SearchResults_ctl0' + strcont + '_certDateLabel1'):
                        datapmi.append(driver.find_element_by_id(
                            'dph_RegistryContent_SearchResults_ctl0' + strcont + '_certDateLabel1').text)
                        cont += 1
                    else:
                        break
                #loop para verificar se o ano do sitem é maior que o ano da planilha base
                for ano in datapmi:
                    anopmi = ano
                    anobase = str(wsbase.cell(row=lin, column=2).value)
                    anobase = anobase[:4]
                    anobase = int(anobase)
                    cont = datapmi.index(ano)+1
                    strcont = str(cont)

                    if anobase <= int(anopmi[len(anopmi) - 4:]):
                        nomepmi = str(driver.find_element_by_id(
                            'dph_RegistryContent_SearchResults_ctl0' + strcont + '_nameLabel1').text).split()
                        if len(nome) >= 3 and len(nomepmi) >= 3:
                            contnome = 0
                            for nb in nome:
                                if len(nb) == 2 or len(nb) == 3:  # Verifica se a quantidade de letras do nome tem 2 ou 3 caracteres
                                    nb = funcoes.verificar_tam_nome(nb, nome)
                                for n in nomepmi:
                                    n = funcoes.verificar_tam_nome(n, nomepmi)
                                    if nb == n:
                                        contnome += 1
                            if contnome >= 3:
                                wsSaida.cell(row=linsaida, column=colNome).value = wsbase.cell(row=lin, column=1).value
                                wsSaida.cell(row=linsaida, column=colData).value = wsbase.cell(row=lin, column=3).value
                                wsSaida.cell(row=linsaida, column=colTurma).value = wsbase.cell(row=lin, column=4).value
                                wsSaida.cell(row=linsaida, column=colNomePMI).value = driver.find_element_by_id(
                                    'dph_RegistryContent_SearchResults_ctl0' + strcont + '_nameLabel1').text
                                wsSaida.cell(row=linsaida, column=colDataPMI).value = (datapmi[cont - 1])
                                wbsaida.save('Template_arquivo_saida.xlsx')
                                linsaida += 1
                                break
                        else:
                            if nome[0] == nomepmi[0]:
                                contnome = 0
                                for nb in nome:
                                    if len(nb) == 2 or len(nb) == 3:  # Verifica se a quantidade de letras do nome tem 2 ou 3 caracteres
                                        nb = funcoes.verificar_tam_nome(nb, nome)
                                    for n in nomepmi:
                                        if len(n) == 2 or len(n) == 3:  # Verifica se a quantidade de letras do nome tem 2 ou 3 caracteres
                                            n = funcoes.verificar_tam_nome(n, nomepmi)
                                        if nb == n:
                                            contnome += 1
                                if contnome >= 2:
                                    wsSaida.cell(row=linsaida, column=colNome).value = wsbase.cell(row=lin, column=1).value
                                    wsSaida.cell(row=linsaida, column=colData).value = wsbase.cell(row=lin, column=2).value
                                    wsSaida.cell(row=linsaida, column=colTurma).value = wsbase.cell(row=lin, column=4).value
                                    wsSaida.cell(row=linsaida, column=colNomePMI).value = driver.find_element_by_id(
                                        'dph_RegistryContent_SearchResults_ctl0' + strcont + '_nameLabel1').text
                                    wsSaida.cell(row=linsaida, column=colDataPMI).value = (datapmi[cont - 1])
                                    wbsaida.save('Template_arquivo_saida.xlsx')
                                    linsaida += 1
                                    break


            c+=1
            sleep(1)
        else:
            c += 1
            sleep(1)

driver.close()
wbsaida.save('Template_arquivo_saida.xlsx')
wbsaida.close()
wbbase.close()
sleep(5)

wbemail = pandas.ExcelFile(dir + '/Backup Admin versão Sponte.xlsx')
plan = wbemail.sheet_names
wsemail = pandas.read_excel(wbemail, wbemail.sheet_names[1])
lista_email = ''
lista_email = wsemail['NOME'].tolist()
wbemail.close()

wb_email = openpyxl.open(dir + '/Backup Admin versão Sponte.xlsx')
sheet = 'Alunos_Dados Pessoais'
ws_email = wb_email[sheet]



colaluno = 1
colemail = 3
colalunoemail = 2
#relsaida = 'C:/Users/Benica.ADMIN-PC/PycharmProjects/ProjectLab_PMI/arquivo_saida/' + arquivo_saida
wbrelsaida = openpyxl.open('Template_arquivo_saida.xlsx')
wsrelsaida = wbrelsaida.active
ultimalinsaida = wsrelsaida.max_row
for linaluno in range(3, ultimalinsaida + 1):
    if wsrelsaida.cell(linaluno, colaluno).value in lista_email:
        linemail = lista_email.index(wsrelsaida.cell(linaluno, colaluno).value) + 2
        if ws_email.cell(linemail, colemail).value is None and ws_email.cell(linemail, colemail+1).value is None:
            wsrelsaida.cell(linaluno, colalunoemail).value = 'Aluno não possui email cadastrado'
        elif ws_email.cell(linemail, colemail).value is not None and ws_email.cell(linemail, colemail + 1).value is not None:
            wsrelsaida.cell(linaluno, colalunoemail).value = str(ws_email.cell(linemail, colemail).value) + '/' + str(ws_email.cell(linemail, colemail + 1).value)
        elif ws_email.cell(linemail, colemail).value is not None and ws_email.cell(linemail, colemail + 1).value is None:
            wsrelsaida.cell(linaluno, colalunoemail).value = ws_email.cell(linemail, colemail).value
        else:
            wsrelsaida.cell(linaluno, colalunoemail).value = ws_email.cell(linemail, colemail + 1).value


wbrelsaida.save('Template_arquivo_saida.xlsx')
wbrelsaida.close()
wb_email.close()
print('Encerrando o Robô...')


arquivo_saida = funcoes.copiar_arquivo('Template_arquivo_saida.xlsx')

sleep(5)
print('Robô Finalizado')
